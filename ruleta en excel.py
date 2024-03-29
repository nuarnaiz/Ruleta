import random
import openpyxl
import os


# Lista de números y colores de la ruleta
colores = ["green"]+["black", "red"] * 18 + ["black", "red"]
numeros=[0,32,15,19,4,21,2,25,17,34,6,27,13,36,11,30,8,23,10,5,24,16,33,1,20,14,31,9,22,18,29,7,28,12,35,3,26]


# Diccionario para mapear números a colores
numero_color = dict(zip(numeros, colores))


# Función para obtener el color de un número
def obtener_color(numero):
  return numero_color.get(numero, "No hay color para ese número")

# Ejemplo de uso
numero = 15
color = obtener_color(numero)
print(f"El número {numero} tiene el color {color}")

# Simular 1000 giros de la ruleta
resultados = []
for _ in range(1000):
  resultado = random.choice(numeros)
  resultados.append(resultado)

# Imprimir resultados con colores
for resultado in resultados:
  color = obtener_color(resultado)


# Crear un nuevo libro de Excel
wb = openpyxl.Workbook()
ws = wb.active

# Escribir encabezados
ws.cell(row=1, column=1).value = "Número"
ws.cell(row=1, column=2).value = "Color"
ws.cell(row=1, column=3).value= "Par o Impar"

# Escribir resultados
for i, resultado in enumerate(resultados, start=2):
  ws.cell(row=i, column=1).value = resultado
  ws.cell(row=i, column=2).value = obtener_color(resultado)
  if resultado%2==0:
    ws.cell(row=i,column=3).value="Par"
  else:
    ws.cell(row=i, column=3).value="Impar"


# Ruta del archivo
ruta_archivo = f"{os.path.join(os.path.expanduser('~'), 'Desktop','master','tfm', 'resultados_ruleta.xlsx')}"

# Guardar el libro de Excel
wb.save(ruta_archivo)

# Imprimir mensaje de éxito
print(f"El archivo se ha guardado en: {ruta_archivo}")